import requests
import random
import asyncio
import openpyxl

wb = openpyxl.load_workbook(filename='word_rus.xlsx')

r = requests.post('http://paraphraser.ru/token/',
                  data={'login': 'annased', 'password': '12345678'})
token = r.json().get('token', '')


def vectorofwrods(line):
    payload = {'c': 'vector',
               'query': line,
               'forms': '0',
               'top': '5',
               'lang': 'ru',
               'format': 'json',
               'scores': '0',
               'token': token}
    r = requests.post('http://paraphraser.ru/api/',
                      data=payload)
    result = r.json()

    if result['code'] == 0:
        value = result['response']
        value = value['1']
        value = value['vector']
        return value
    else:
        print('Error:', result['msg'])


def synonyms(line):
    payload = {'c': 'syns',
               'query': line,
               'forms': '0',
               'scores': '0',
               'top': '5',
               'lang': 'ru',
               'format': 'json',
               'token': token}

    r = requests.post('http://paraphraser.ru/api/',
                      data=payload)
    result = r.json()

    if result['code'] == 0:
        value = result['response']
        value = value['1']
        value = value['syns']
        return value
    else:
        print('Error:', result['msg'])


def hyps(line):
    payload = {'c': 'hyp',
               'query': line,
               'forms': '0',
               'scores': '0',
               'top': '5',
               'lang': 'ru',
               'format': 'json',
               'token': token}

    r = requests.post('http://paraphraser.ru/api/',
                      data=payload)
    result = r.json()

    if result['code'] == 0:
        value = result['response']
        value = value['1']
        value1 = value['hyponym']
        value1 = value1[line]
        value2 = value['hypernym']
        value2 = value2[line]
        if value1 is None:
            value1 = list()
        if value2 is None:
            value2 = list()
        return list(value1, value2)
    else:
        print('Error:', result['msg'])


def generate_5_words(line):
    a = vectorofwrods(line)
    b = synonyms(line)
    c = list()
    d = list()
    result = set()
    if len(b) > 0:
        result.add(b[random.randint(0, len(b)-1)])
    if len(c) > 0:
        result.add(c[random.randint(0, len(c)-1)])
    if len(d) > 0:
        result.add(d[random.randint(0, len(d)-1)])
    b = b + c + d + a
    if line in result:
        result.remove(line)
    for i in a:
        if random.randint(0, 1) and len(result) < 5:
            result.add(i)
    if len(result) < 5:
        for j in b:
            if j not in result and len(result) < 5:
                result.add(j)
                if line in result:
                    result.remove(line)
    return result


def Generate_Words(difficulty, results=tuple()):
    if difficulty == 1:
        sheet = wb['easy']
        size = len(sheet['A'])
    else:
        sheet = wb['medium']
        size = len(sheet['A'])
    row = random.randint(1, size)
    current = sheet.cell(row=row, column=1).value
    values = generate_5_words(current)
    while len(values) < 5:
        sheet.delete_rows(row)
        row = random.randint(1, size)
        current = sheet.cell(row=row, column=1).value
        values = generate_5_words(current)
    results = [current, values]
    return [current, *values]
